"""
Athena School Memory Indexer

Indexes supported files from the School folder into
Athena's semantic memory using Kev + ChromaDB.

Only new or changed files are re-indexed.
"""

from pathlib import Path
import hashlib
import json
import csv
from openpyxl import load_workbook

from pypdf import PdfReader
from docx import Document

from memory.memory_store import (
    store_memory,
    collection,
)


SCHOOL_PATH = Path(
    "/Users/daneanderson/Desktop/School"
)

INDEX_FILE = (
    Path(__file__).parent
    / "school_index.json"
)

SUPPORTED_EXTENSIONS = {
    ".txt",
    ".md",
    ".pdf",
    ".docx",
    ".xlsx",
    ".xlsm",
    ".csv",
}

CHUNK_SIZE = 1500
CHUNK_OVERLAP = 200

def get_memory_scope(path: Path):
    """
    Determine Athena memory scope from folder structure.
    """

    folder_text = " ".join(
        part.lower()
        for part in path.parts
    )

    if "csci" in folder_text:
        return "computer_science"

    if (
        "appm" in folder_text
        or "calc" in folder_text
        or "mit 18.01" in folder_text
    ):
        return "mathematics"

    if "phil" in folder_text:
        return "humanities"

    if "internship" in folder_text:
        return "internships"

    if "transfer" in folder_text:
        return "transfer"

    if "project" in folder_text:
        return "projects"

    return "school"


def file_hash(path: Path):
    """
    Create a hash so Athena can detect
    whether a file changed.
    """

    hasher = hashlib.sha256()

    with open(path, "rb") as file:
        while chunk := file.read(1024 * 1024):
            hasher.update(chunk)

    return hasher.hexdigest()


def load_index():
    if not INDEX_FILE.exists():
        return {}

    return json.loads(
        INDEX_FILE.read_text(
            encoding="utf-8"
        )
    )


def save_index(index):
    INDEX_FILE.write_text(
        json.dumps(
            index,
            indent=2,
        ),
        encoding="utf-8",
    )


def extract_text(path: Path):
    """
    Extract text from supported school files.
    """

    extension = path.suffix.lower()

    if extension in {
        ".txt",
        ".md",
    }:
        return path.read_text(
            encoding="utf-8",
            errors="ignore",
        )

    if extension == ".pdf":

        reader = PdfReader(path)

        pages = []

        for page in reader.pages:
            text = page.extract_text()

            if text:
                pages.append(text)

        return "\n\n".join(pages)

    if extension == ".docx":

        try:
            document = Document(path)

            return "\n".join(
                paragraph.text
                for paragraph
                in document.paragraphs
            )

        except Exception as error:
            print(
                f"Skipping unreadable DOCX: "
                f"{path.name} ({error})"
            )
            return ""

        return ""
    
    if extension in {
        ".xlsx",
        ".xlsm",
    }:
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
            )

            output = []

            for sheet in workbook.worksheets:

                output.append(
                    f"\nSHEET: {sheet.title}\n"
                )

                for row in sheet.iter_rows(
                    values_only=True
                ):
                    values = [
                        str(value)
                        if value is not None
                        else ""
                        for value in row
                    ]

                    if any(values):
                        output.append(
                            "\t".join(values)
                        )

            workbook.close()

            return "\n".join(output)

        except Exception as error:
            print(
                f"Skipping unreadable spreadsheet: "
                f"{path.name} ({error})"
            )
            return ""

    if extension == ".csv":
        try:
            rows = []

            with open(
                path,
                "r",
                encoding="utf-8-sig",
                errors="ignore",
                newline="",
            ) as file:

                reader = csv.reader(file)

                for row in reader:
                    rows.append(
                        "\t".join(row)
                    )

            return "\n".join(rows)

        except Exception as error:
            print(
                f"Skipping unreadable CSV: "
                f"{path.name} ({error})"
            )
            return ""


def chunk_text(text):
    """
    Split text into overlapping chunks.
    """

    chunks = []

    start = 0

    while start < len(text):

        end = start + CHUNK_SIZE

        chunk = text[start:end].strip()

        if chunk:
            chunks.append(chunk)

        start += (
            CHUNK_SIZE
            - CHUNK_OVERLAP
        )

    return chunks


def index_file(
    path: Path,
    index,
):
    """
    Embed one file and store its chunks.
    """

    source = str(path)

    current_hash = file_hash(path)

    previous = index.get(source)

    if (
        previous
        and previous["hash"]
        == current_hash
    ):
        print(
            f"Skipping unchanged: "
            f"{path.name}"
        )
        return

    if previous:
        old_ids = previous.get(
            "chunk_ids",
            [],
        )

        if old_ids:
            collection.delete(
                ids=old_ids
            )

    text = extract_text(path)

    if not text.strip():
        print(
            f"No readable text: "
            f"{path.name}"
        )
        return

    chunks = chunk_text(text)

    chunk_ids = []

    for number, chunk in enumerate(
        chunks
    ):

        memory_id = (
            hashlib.sha256(
                (
                    source
                    + current_hash
                    + str(number)
                ).encode()
            ).hexdigest()
        )

        metadata = {
            "collection": "school",
            "scope": get_memory_scope(path),
            "source": source,
            "filename": path.name,
            "extension": (
                path.suffix.lower()
            ),
            "chunk": number,
        }

        store_memory(
            memory_id=memory_id,
            text=chunk,
            metadata=metadata,
        )

        chunk_ids.append(
            memory_id
        )

    index[source] = {
        "hash": current_hash,
        "chunk_ids": chunk_ids,
    }

    print(
        f"Indexed: {path.name} "
        f"({len(chunks)} chunks)"
    )


def index_school_folder():
    """
    Index all supported files in School.
    """

    index = load_index()

    files = [
        path
        for path
        in SCHOOL_PATH.rglob("*")
        if (
            path.is_file()
            and path.suffix.lower()
            in SUPPORTED_EXTENSIONS
            and not path.name.startswith("~$")
        )
    ]

    print(
        f"Found {len(files)} "
        f"supported school files."
    )

    for path in files:
        index_file(
            path,
            index,
        )

        save_index(index)

    print(
        "School memory indexing complete."
    )


if __name__ == "__main__":
    index_school_folder()